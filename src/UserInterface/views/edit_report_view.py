"""Edit Report View - In-app data editor with database integration and versioning."""

import os
import shutil
import platform
import subprocess
from typing import Optional, Dict, Any, List
from datetime import datetime

import customtkinter as ctk
from tkinter import messagebox, ttk
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from UserInterface.services.data_validator import DataValidator
from UserInterface.services.database_service import DatabaseService
from excel_manager import ExcelManager
from powerpoint_generator import PowerPointGenerator


class EditReportView:
    """View for editing existing report data with database integration."""

    def __init__(self, parent: ctk.CTk, controller):
        self.parent = parent
        self.controller = controller
        self.validator = DataValidator()
        self.equipment_map: Dict[str, Any] = {}
        self.work_name: str = ""
        self.excel_path: str = ""
        self.file_type: str = "excel"
        self.db: Optional[Session] = None
        self.user_id: Optional[int] = None
        self.work_id: Optional[int] = None
        self.project_root = os.path.abspath(
            os.path.join(os.path.dirname(__file__), "..", "..", "..")
        )
        
        # Treeview components
        self.tree = None
        self.tree_frame = None
        self.entry_widget = None
        self.editing_item = None
        self.editing_column = None
        
        # Data
        self.table_data = []  # List of dictionaries for each row
        self.table_headers = []
        self.editable_columns = []  # Column indices that are editable
        
        # Data mapping
        self.row_to_equipment_map = {}  # Maps tree row index to (equipment_no, component_name)

    def show(self, edit_context: Optional[Dict[str, Any]] = None) -> None:
        """Display the edit report interface."""
        if edit_context is None:
            edit_context = getattr(self.controller, 'edit_context', None)
        
        if not edit_context:
            messagebox.showerror("Error", "No report selected for editing.")
            if hasattr(self.controller, 'show'):
                self.controller.show()
            return
        
        self.work_name = edit_context.get('work_name', '')
        file_path = edit_context.get('file_path', '')
        self.file_type = edit_context.get('file_type', 'excel')
        self.db = edit_context.get('db')
        self.user_id = edit_context.get('user_id')
        
        # Get work_id from database
        if self.db and self.work_name:
            try:
                from AutoRBI_Database.database.models.work import Work
                work = self.db.query(Work).filter(Work.work_name == self.work_name).first()
                if work:
                    self.work_id = work.work_id
            except Exception as e:
                print(f"Error fetching work_id: {e}")
        
        if not self.work_name or not file_path:
            messagebox.showerror("Error", "Invalid report context.")
            if hasattr(self.controller, 'show'):
                self.controller.show()
            return
        
        # Handle different file types
        if self.file_type == 'powerpoint':
            self._edit_powerpoint(file_path)
        else:  # Excel
            self.excel_path = file_path
            self._load_equipment_data()

    def _edit_powerpoint(self, file_path: str) -> None:
        """Handle PowerPoint file editing (open with default application)."""
        try:
            if not os.path.exists(file_path):
                messagebox.showerror("File Not Found", f"File does not exist:\n{file_path}")
                self.controller.show()
                return
            
            # Log action to database
            if self.db and self.work_id and self.user_id:
                DatabaseService.log_work_history(
                    self.db,
                    self.work_id,
                    self.user_id,
                    action_type="edit_report",
                    description="Opened PowerPoint file for editing"
                )
            
            # Open the PowerPoint file
            if platform.system() == 'Windows':
                os.startfile(file_path)
            elif platform.system() == 'Darwin':
                subprocess.Popen(['open', file_path])
            else:
                subprocess.Popen(['xdg-open', file_path])
            
            messagebox.showinfo(
                "PowerPoint Opened",
                "The PowerPoint file has been opened in your default application.\n\n"
                "Edit the file as needed and save your changes.\n"
                "Then return to the Report Menu."
            )
            self.controller.show()
            
        except Exception as e:
            messagebox.showerror("Error", f"Could not open PowerPoint file:\n{e}")
            self.controller.show()

    def _load_equipment_data(self) -> None:
        """Load equipment data from Excel file."""
        try:
            excel_manager = ExcelManager(self.excel_path)
            self.equipment_map = excel_manager.read_masterfile()
            
            if not self.equipment_map:
                messagebox.showerror("Error", "Failed to load equipment data from Excel.")
                self.controller.show()
                return
            
            # Prepare data for Treeview
            self._prepare_table_data()
            
            # Build UI
            self._build_ui()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel data:\n{e}")
            self.controller.show()

    def _prepare_table_data(self) -> None:
        """Convert equipment_map to list of dictionaries for Treeview."""
        self.table_headers = [
            "No.", "Equipment No.", "PMT No.", "Equipment Description", 
            "Component", "Phase", "Fluid", "Material Type", "Spec", 
            "Grade", "Insulation", "Design Temp", "Design Pressure", 
            "Operating Temp", "Operating Pressure"
        ]
        
        # Define which columns are editable by column name
        self.editable_columns = [
            "Fluid", "Material Type", "Spec", "Grade", "Insulation",
            "Design Temp", "Design Pressure", "Operating Temp", "Operating Pressure"
        ]
        
        # Create a mapping from column name to index for quick lookup
        self.column_name_to_index = {name: idx for idx, name in enumerate(self.table_headers)}
        
        self.table_data = []  # This will store dictionaries
        self.row_to_equipment_map = {}
        row_counter = 1
        
        for eq_no, equipment in self.equipment_map.items():
            for component in equipment.components:
                # Create row dictionary with all field names as keys
                row_dict = {
                    'No.': str(row_counter),
                    'Equipment No.': self._safe_str(eq_no),
                    'PMT No.': self._safe_str(equipment.pmt_number),
                    'Equipment Description': self._safe_str(equipment.equipment_description),
                    'Component': self._safe_str(component.component_name),
                    'Phase': self._safe_str(getattr(component, 'phase', '')),
                    'Fluid': self._safe_str(self._get_component_value_safe(component, 'fluid')),
                    'Material Type': self._safe_str(self._get_component_value_safe(component, 'material_type')),
                    'Spec': self._safe_str(self._get_component_value_safe(component, 'spec')),
                    'Grade': self._safe_str(self._get_component_value_safe(component, 'grade')),
                    'Insulation': self._safe_str(self._get_component_value_safe(component, 'insulation')),
                    'Design Temp': self._safe_str(self._get_component_value_safe(component, 'design_temp')),
                    'Design Pressure': self._safe_str(self._get_component_value_safe(component, 'design_pressure')),
                    'Operating Temp': self._safe_str(self._get_component_value_safe(component, 'operating_temp')),
                    'Operating Pressure': self._safe_str(self._get_component_value_safe(component, 'operating_pressure'))
                }
                
                self.table_data.append(row_dict)
                self.row_to_equipment_map[len(self.table_data) - 1] = (eq_no, component.component_name)
                row_counter += 1

    def _safe_str(self, value) -> str:
        """Safely convert any value to string."""
        if value is None:
            return ""
        try:
            return str(value)
        except:
            return ""

    def _get_component_value_safe(self, component, key: str) -> str:
        """Safely get a value from component with multiple fallback methods."""
        try:
            # Method 1: Try to get from existing_data
            if hasattr(component, 'existing_data') and component.existing_data:
                value = component.existing_data.get(key)
                if value is not None:
                    return str(value)
            
            # Method 2: Try get_existing_data_value method
            if hasattr(component, 'get_existing_data_value'):
                value = component.get_existing_data_value(key)
                if value is not None:
                    return str(value)
            
            # Method 3: Try direct attribute
            if hasattr(component, key):
                value = getattr(component, key)
                if value is not None:
                    return str(value)
            
            # Return empty string if nothing found
            return ""
            
        except Exception:
            return ""

    def _build_ui(self) -> None:
        """Build the edit interface with Treeview."""
        for widget in self.parent.winfo_children():
            widget.destroy()

        root_frame = ctk.CTkFrame(self.parent, corner_radius=0, fg_color="transparent")
        root_frame.pack(expand=True, fill="both", padx=32, pady=24)

        root_frame.grid_rowconfigure(1, weight=1)
        root_frame.grid_columnconfigure(0, weight=1)

        # Header
        header = ctk.CTkFrame(root_frame, fg_color="transparent")
        header.grid(row=0, column=0, sticky="ew")
        header.grid_columnconfigure(1, weight=1)

        back_btn = ctk.CTkButton(
            header,
            text="← Back to Reports",
            command=self._confirm_back,
            width=160,
            height=32,
            font=("Segoe UI", 10),
            fg_color="transparent",
            text_color=("gray20", "gray90"),
            hover_color=("gray85", "gray30"),
        )
        back_btn.grid(row=0, column=0, sticky="w")

        title_label = ctk.CTkLabel(
            header,
            text="AutoRBI - Edit Report",
            font=("Segoe UI", 20, "bold"),
        )
        title_label.grid(row=0, column=1)

        # Action buttons
        action_frame = ctk.CTkFrame(header, fg_color="transparent")
        action_frame.grid(row=0, column=2, sticky="e")

        validate_btn = ctk.CTkButton(
            action_frame,
            text="🔍 Validate",
            command=self._validate_data,
            width=100,
            height=32,
            font=("Segoe UI", 10),
            fg_color=("orange", "darkorange"),
        )
        validate_btn.pack(side="left", padx=(0, 8))

        save_btn = ctk.CTkButton(
            action_frame,
            text="💾 Save & Regenerate",
            command=self._save_and_regenerate,
            width=160,
            height=32,
            font=("Segoe UI", 10),
            fg_color=("green", "darkgreen"),
        )
        save_btn.pack(side="left")

        # Main content
        main_frame = ctk.CTkFrame(
            root_frame,
            corner_radius=18,
            border_width=1,
            border_color=("gray80", "gray25"),
        )
        main_frame.grid(row=1, column=0, sticky="nsew", pady=(12, 0))

        main_frame.grid_rowconfigure(2, weight=1)
        main_frame.grid_columnconfigure(0, weight=1)

        # Info section
        info_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        info_frame.grid(row=0, column=0, sticky="ew", padx=24, pady=(18, 6))

        work_label = ctk.CTkLabel(
            info_frame,
            text=f"Editing: {self.work_name} ({len(self.table_data)} rows)",
            font=("Segoe UI", 16, "bold"),
            anchor="w"
        )
        work_label.pack(anchor="w")

        instruction_label = ctk.CTkLabel(
            info_frame,
            text="Double-click on editable cells to modify values. Columns Fluid through Operating Pressure are editable.",
            font=("Segoe UI", 11),
            text_color=("gray25", "gray80"),
            anchor="w"
        )
        instruction_label.pack(anchor="w", pady=(4, 0))

        # Table container
        self.tree_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        self.tree_frame.grid(row=2, column=0, sticky="nsew", padx=24, pady=(12, 24))
        self.tree_frame.grid_rowconfigure(0, weight=1)
        self.tree_frame.grid_columnconfigure(0, weight=1)
        
        # Create Treeview
        self._create_treeview()

    def _create_treeview(self) -> None:
        """Create ttk.Treeview widget for displaying and editing data."""
        # Create a frame to hold the treeview and scrollbars
        container = ctk.CTkFrame(self.tree_frame, fg_color="transparent")
        container.pack(expand=True, fill="both")
        
        # Create vertical scrollbar
        v_scrollbar = ttk.Scrollbar(container, orient="vertical")
        v_scrollbar.pack(side="right", fill="y")
        
        # Create horizontal scrollbar
        h_scrollbar = ttk.Scrollbar(container, orient="horizontal")
        h_scrollbar.pack(side="bottom", fill="x")
        
        # Create the Treeview with custom style
        style = ttk.Style()
        style.configure("Custom.Treeview", 
                       background="white",
                       foreground="black",
                       fieldbackground="white",
                       rowheight=25)
        style.configure("Custom.Treeview.Heading",
                       background=("#e0e0e0", "#2d2d2d"),
                       foreground=("black", "white"),
                       relief="flat")
        style.map("Custom.Treeview.Heading",
                 background=[('active', '#d0d0d0')])
        
        # Add "_index" to columns if we need to track row indices
        all_columns = self.table_headers[:]  # Copy the headers
        
        self.tree = ttk.Treeview(
            container,
            columns=all_columns,
            show="headings",
            style="Custom.Treeview",
            yscrollcommand=v_scrollbar.set,
            xscrollcommand=h_scrollbar.set,
            height=20,
            selectmode="browse"  # Only select one row at a time
        )
        
        # Configure scrollbars
        v_scrollbar.config(command=self.tree.yview)
        h_scrollbar.config(command=self.tree.xview)
        
        # Pack the treeview
        self.tree.pack(side="left", expand=True, fill="both")
        
        # Define column headings and widths
        column_widths = {
            "No.": 40,
            "Equipment No.": 100,
            "PMT No.": 80,
            "Equipment Description": 150,
            "Component": 100,
            "Phase": 60,
            "Fluid": 80,
            "Material Type": 90,
            "Spec": 80,
            "Grade": 70,
            "Insulation": 80,
            "Design Temp": 90,
            "Design Pressure": 100,
            "Operating Temp": 100,
            "Operating Pressure": 110
        }
        
        # Configure columns - only show the actual data columns
        for header in self.table_headers:
            self.tree.heading(header, text=header)
            self.tree.column(header, width=column_widths.get(header, 100), anchor="w")
        
        # Insert data and store mapping in a separate dictionary
        self.tree_item_to_data_index = {}  # Maps tree item ID to data index
        
        for i, row_dict in enumerate(self.table_data):
            values = [row_dict.get(header, "") for header in self.table_headers]
            item_id = self.tree.insert("", "end", values=values)
            
            # Store mapping from tree item ID to data index
            self.tree_item_to_data_index[item_id] = i
        
        # Make the treeview editable
        self._make_treeview_editable()

    def _make_treeview_editable(self) -> None:
        """Make specific columns in Treeview editable with double-click."""
        # Create entry widget for editing (initially hidden)
        self.entry_widget = ttk.Entry(self.tree_frame, font=("Segoe UI", 10))
        self.entry_widget.bind("<Return>", self._save_edit)
        self.entry_widget.bind("<Escape>", self._cancel_edit)
        self.entry_widget.bind("<FocusOut>", self._save_edit)
        
        # Bind double-click to start editing
        self.tree.bind("<Double-1>", self._on_double_click)
        
    def _on_double_click(self, event) -> None:
        """Handle double-click to edit cell."""
        # Identify which cell was clicked
        row_id = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        
        if not row_id or column == "#0":  # #0 is the tree column
            return
        
        # Get column index and name
        col_index = int(column[1:]) - 1
        col_name = self.table_headers[col_index]
        
        # Only allow editing of editable columns
        if col_name in self.editable_columns:
            # Get the cell coordinates
            x, y, width, height = self.tree.bbox(row_id, column)
            
            # Get current value
            current_value = self.tree.set(row_id, col_name)
            
            # Position and configure the entry widget
            self.entry_widget.place(x=x, y=y, width=width, height=height)
            self.entry_widget.delete(0, "end")
            self.entry_widget.insert(0, current_value)
            self.entry_widget.select_range(0, "end")
            self.entry_widget.focus()
            
            # Store editing state
            self.editing_item = row_id
            self.editing_column = col_name
            
    def _save_edit(self, event=None) -> None:
        """Save the edited value."""
        if not self.editing_item or not self.editing_column:
            return
        
        # Get the new value
        new_value = self.entry_widget.get()
        
        # Get the old value
        old_value = self.tree.set(self.editing_item, self.editing_column)
        
        if new_value != old_value:
            # Update the treeview
            self.tree.set(self.editing_item, self.editing_column, new_value)
            
            # Get the data index for this row using our mapping
            if self.editing_item in self.tree_item_to_data_index:
                data_index = self.tree_item_to_data_index[self.editing_item]
                
                # Update table_data
                if 0 <= data_index < len(self.table_data):
                    # Update the dictionary
                    self.table_data[data_index][self.editing_column] = new_value
                    
                    # Update equipment_map
                    if data_index in self.row_to_equipment_map:
                        eq_no, component_name = self.row_to_equipment_map[data_index]
                        if eq_no in self.equipment_map:
                            equipment = self.equipment_map[eq_no]
                            component = equipment.get_component(component_name)
                            if component:
                                # Map column name to field name in existing_data
                                field_name_map = {
                                    'Fluid': 'fluid',
                                    'Material Type': 'material_type',
                                    'Spec': 'spec',
                                    'Grade': 'grade',
                                    'Insulation': 'insulation',
                                    'Design Temp': 'design_temp',
                                    'Design Pressure': 'design_pressure',
                                    'Operating Temp': 'operating_temp',
                                    'Operating Pressure': 'operating_pressure'
                                }
                                
                                if self.editing_column in field_name_map:
                                    field_name = field_name_map[self.editing_column]
                                    component.set_existing_data_value(field_name, new_value)
        
        # Hide the entry widget
        self.entry_widget.place_forget()
        self.editing_item = None
        self.editing_column = None
        
    def _cancel_edit(self, event=None) -> None:
        """Cancel editing."""
        self.entry_widget.place_forget()
        self.editing_item = None
        self.editing_column = None
        self.editing_col_index = None

    def _validate_data(self) -> None:
        """Validate the edited data."""
        if not self.table_data:
            return
        
        # Simple validation - check for required fields
        errors = []
        
        for row_idx, row_dict in enumerate(self.table_data):
            # Check fluid is not empty
            if not str(row_dict.get('Fluid', '')).strip():
                errors.append(f"Row {row_idx + 1}: Fluid is required")
            
            # Check material type is not empty
            if not str(row_dict.get('Material Type', '')).strip():
                errors.append(f"Row {row_idx + 1}: Material Type is required")
        
        if errors:
            error_msg = "❌ Validation errors found:\n\n" + "\n".join(errors[:10])  # Show first 10 errors
            if len(errors) > 10:
                error_msg += f"\n\n... and {len(errors) - 10} more errors."
            messagebox.showerror("Validation Failed", error_msg)
        else:
            messagebox.showinfo("Validation Success", "✅ All data is valid and ready to save!")

    def _save_and_regenerate(self) -> None:
        """Save edited data and regenerate Excel and PowerPoint."""
        response = messagebox.askyesno(
            "Confirm Save",
            "This will create a new version of the report.\n\n"
            "New files will be generated:\n"
            "• Updated Excel file\n"
            "• Updated PowerPoint presentation\n\n"
            "Continue?"
        )
        
        if not response:
            return
        
        progress_window = self._create_progress_window()
        
        try:
            # Ensure all edits are saved
            if self.editing_item:
                self._save_edit()
            
            # Update equipment_map with all edited data
            self._update_equipment_map_from_table()
            
            progress_window.update_progress(0.2, "Updating database...")
            self._save_to_database()
            
            progress_window.update_progress(0.4, "Saving Excel file...")
            excel_path = self._save_edited_excel()
            
            if not excel_path:
                raise Exception("Failed to save Excel file")
            
            progress_window.update_progress(0.6, "Generating PowerPoint...")
            ppt_path = self._generate_powerpoint()
            
            if not ppt_path:
                raise Exception("Failed to generate PowerPoint")
            
            progress_window.update_progress(0.8, "Updating work records...")
            self._update_work_paths(excel_path, ppt_path)
            
            progress_window.update_progress(0.95, "Logging action...")
            self._log_edit_action()
            
            progress_window.update_progress(1.0, "Complete!")
            progress_window.close()
            
            messagebox.showinfo(
                "Success",
                f"✅ New version created successfully!\n\n"
                f"Excel: {os.path.basename(excel_path)}\n"
                f"PowerPoint: {os.path.basename(ppt_path)}\n\n"
                f"You can now view the new version in the Report Menu."
            )
            
            self.controller.show()
            
        except Exception as e:
            progress_window.close()
            messagebox.showerror("Error", f"Failed to save and regenerate:\n{e}")

    def _update_equipment_map_from_table(self) -> None:
        """Update equipment map with edited data from table."""
        try:
            for row_idx, (eq_no, component_name) in self.row_to_equipment_map.items():
                if eq_no in self.equipment_map and row_idx < len(self.table_data):
                    equipment = self.equipment_map[eq_no]
                    component = equipment.get_component(component_name)
                    
                    if component:
                        row_dict = self.table_data[row_idx]
                        
                        updates = {
                            'fluid': str(row_dict.get('Fluid', '')).strip(),
                            'material_type': str(row_dict.get('Material Type', '')).strip(),
                            'spec': str(row_dict.get('Spec', '')).strip(),
                            'grade': str(row_dict.get('Grade', '')).strip(),
                            'insulation': self.validator.normalize_insulation(str(row_dict.get('Insulation', '')).strip()),
                            'design_temp': self.validator.normalize_temperature(str(row_dict.get('Design Temp', '')).strip()),
                            'design_pressure': self.validator.normalize_pressure(str(row_dict.get('Design Pressure', '')).strip()),
                            'operating_temp': self.validator.normalize_temperature(str(row_dict.get('Operating Temp', '')).strip()),
                            'operating_pressure': self.validator.normalize_pressure(str(row_dict.get('Operating Pressure', '')).strip())
                        }
                        
                        component.update_existing_data(updates)
                        
        except Exception as e:
            print(f"Error updating equipment map from table: {e}")

    def _save_to_database(self) -> None:
        """Save updated equipment and components to database."""
        if not self.db or not self.work_id or not self.user_id:
            print("Warning: Database context not available, skipping database save")
            return
        
        try:
            # Batch save equipment with updated components
            success_count, failure_count = DatabaseService.batch_save_equipment(
                self.db,
                self.work_id,
                self.user_id,
                self.equipment_map,
                {}  # drawing_paths - not needed for editing
            )
            
            print(f"Database save: {success_count} successful, {failure_count} failed")
            
        except Exception as e:
            print(f"Error saving to database: {e}")

    def _save_edited_excel(self) -> Optional[str]:
        """Save edited data to new Excel file in 'edited' folder."""
        try:
            work_dir = os.path.join(self.project_root, "src", "output_files", self.work_name)
            edited_dir = os.path.join(work_dir, "excel", "edited")
            os.makedirs(edited_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.work_name}_edited_{timestamp}.xlsx"
            output_path = os.path.join(edited_dir, filename)
            
            # Copy original file
            shutil.copy2(self.excel_path, output_path)
            
            # Update with new data
            wb = load_workbook(output_path)
            
            if 'Masterfile' not in wb.sheetnames:
                print(f"Warning: 'Masterfile' sheet not found in {output_path}")
                wb.close()
                return output_path
            
            ws = wb['Masterfile']
            
            for equipment in self.equipment_map.values():
                for component in equipment.components:
                    row_index = getattr(component, 'row_index', None)
                    if not row_index:
                        continue
                    
                    row = row_index
                    
                    fluid = self._safe_get_component_value(component, 'fluid', '')
                    material = self._safe_get_component_value(component, 'material_type', '')
                    spec = self._safe_get_component_value(component, 'spec', '')
                    grade = self._safe_get_component_value(component, 'grade', '')
                    insulation = self._safe_get_component_value(component, 'insulation', '')
                    design_temp = self._safe_get_component_value(component, 'design_temp', '')
                    design_pressure = self._safe_get_component_value(component, 'design_pressure', '')
                    operating_temp = self._safe_get_component_value(component, 'operating_temp', '')
                    operating_pressure = self._safe_get_component_value(component, 'operating_pressure', '')
                    
                    ws[f'G{row}'] = fluid
                    ws[f'H{row}'] = material
                    ws[f'I{row}'] = spec
                    ws[f'J{row}'] = grade
                    ws[f'K{row}'] = insulation
                    ws[f'L{row}'] = design_temp
                    ws[f'M{row}'] = design_pressure
                    ws[f'N{row}'] = operating_temp
                    ws[f'O{row}'] = operating_pressure
            
            wb.save(output_path)
            wb.close()
            
            return output_path
            
        except Exception as e:
            print(f"Error saving edited Excel: {e}")
            return None

    def _safe_get_component_value(self, component, key: str, default: str = "") -> str:
        """Safely get a value from component's existing_data."""
        try:
            if hasattr(component, 'existing_data'):
                return component.existing_data.get(key, default)
            elif hasattr(component, 'get_existing_data_value'):
                value = component.get_existing_data_value(key)
                return value if value else default
            else:
                return default
        except Exception:
            return default

    def _generate_powerpoint(self) -> Optional[str]:
        """Generate new PowerPoint from edited data."""
        try:
            work_dir = os.path.join(self.project_root, "src", "output_files", self.work_name)
            edited_dir = os.path.join(work_dir, "powerpoint", "edited")
            os.makedirs(edited_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.work_name}_edited_{timestamp}.pptx"
            output_path = os.path.join(edited_dir, filename)
            
            template_path = os.path.join(
                self.project_root, "CaseStudy1Resources", "Inspection Plan Template.pptx"
            )
            
            if not os.path.exists(template_path):
                print(f"Template not found: {template_path}")
                return None
            
            generator = PowerPointGenerator(template_path)
            success = generator.generate_from_equipment_map(self.equipment_map, output_path)
            
            return output_path if success else None
            
        except Exception as e:
            print(f"Error generating PowerPoint: {e}")
            return None

    def _update_work_paths(self, excel_path: str, ppt_path: str) -> None:
        """Update the Work model with new file paths."""
        if not self.db or not self.work_id:
            return
        
        try:
            from AutoRBI_Database.database.models.work import Work
            
            work = self.db.query(Work).filter(Work.work_id == self.work_id).first()
            if work:
                work.excel_path = excel_path
                work.ppt_path = ppt_path
                self.db.commit()
                print(f"Updated Work {self.work_id} paths")
                
        except Exception as e:
            print(f"Error updating work paths: {e}")

    def _log_edit_action(self) -> None:
        """Log the edit action to database."""
        if not self.db or not self.work_id or not self.user_id:
            return
        
        try:
            equipment_count = len(self.equipment_map)
            DatabaseService.log_work_history(
                self.db,
                self.work_id,
                self.user_id,
                action_type="edit_report",
                description=f"Edited and regenerated report with {equipment_count} equipment items"
            )
        except Exception as e:
            print(f"Error logging edit action: {e}")

    def _confirm_back(self) -> None:
        """Confirm before going back (unsaved changes warning)."""
        response = messagebox.askyesno(
            "Confirm",
            "Are you sure you want to go back?\n\nAny unsaved changes will be lost."
        )
        
        if response:
            self.controller.show()

    def _create_progress_window(self):
        """Create a progress window for save operations."""
        return ProgressWindow(self.parent)


class ProgressWindow:
    """Simple progress window for long operations."""
    
    def __init__(self, parent):
        self.window = ctk.CTkToplevel(parent)
        self.window.title("Processing")
        self.window.geometry("400x150")
        self.window.transient(parent)
        self.window.grab_set()
        
        self.window.update_idletasks()
        x = (self.window.winfo_screenwidth() // 2) - 200
        y = (self.window.winfo_screenheight() // 2) - 75
        self.window.geometry(f'400x150+{x}+{y}')
        
        self.label = ctk.CTkLabel(
            self.window,
            text="Processing...",
            font=("Segoe UI", 12)
        )
        self.label.pack(pady=(30, 10))
        
        self.progress = ctk.CTkProgressBar(self.window, width=350)
        self.progress.pack(pady=10)
        self.progress.set(0)
        
        self.status = ctk.CTkLabel(
            self.window,
            text="",
            font=("Segoe UI", 10),
            text_color=("gray60", "gray70")
        )
        self.status.pack(pady=5)
    
    def update_progress(self, value: float, text: str = ""):
        """Update progress bar and text."""
        self.progress.set(value)
        self.status.configure(text=text)
        self.window.update()
    
    def close(self):
        """Close the progress window."""
        self.window.destroy()