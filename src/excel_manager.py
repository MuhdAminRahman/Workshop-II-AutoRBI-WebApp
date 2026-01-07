# excel_manager.py
from datetime import datetime
from openpyxl import load_workbook
import logging
from typing import List, Dict, Optional
from models import Equipment, Component
import os
import re



logger = logging.getLogger(__name__)

class ExcelManager:
    def __init__(self, file_path: str, log_callback=None):
        self.file_path = file_path
        self.wb = None
        self._is_loaded = False
        self.equipment_map: Dict[str, Equipment] = {}
        self.default_path = "src/output_files"
        self.log_callback = log_callback
        self.basefilename = None
    
    def load_workbook(self):
        """Explicitly load the workbook when needed"""
        if not self._is_loaded:
            self.wb = load_workbook(self.file_path)
            self._is_loaded = True
            self.log_info(f"Loaded workbook: {self.file_path}")

    def close_workbook(self):
        """Explicitly close the workbook"""
        if self._is_loaded and self.wb:
            self.wb.close()
            self._is_loaded = False
            self.log_info("Closed workbook")


    def read_masterfile(self) -> Dict[str, Equipment]:
        """
        Read the existing equipment structure from the Masterfile sheet
        Returns a dictionary of Equipment objects with equipment_number as key
        """
        self.load_workbook()
        equipment_map = {}
        ws = self.wb['Masterfile']
        
        current_row = 7
        current_equipment = None
        
        while current_row <= ws.max_row and current_row <= 50:  # Limit to first 50 rows
            equipment_number = self._get_cell_value(ws, f'B{current_row}')
            component_name = self._get_cell_value(ws, f'E{current_row}')
            
            if equipment_number and equipment_number not in ['EQUIPMENT NO.', '']:
                # New equipment found
                if current_equipment and current_equipment.components:
                    # Add to dictionary using equipment_number as key
                    equipment_map[current_equipment.equipment_number] = current_equipment
                
                current_equipment = Equipment(
                    equipment_number=equipment_number,
                    pmt_number=self._get_cell_value(ws, f'C{current_row}'),
                    equipment_description=self._get_cell_value(ws, f'D{current_row}'),
                    row_index=current_row
                )
                self.log_info(f"Found equipment: {equipment_number}")
            
            if current_equipment and component_name and component_name not in ['COMPONENTS', '']:
                # Create and add component information
                component_data = Component(
                    component_name=component_name,
                    phase=self._get_cell_value(ws, f'F{current_row}'),
                    row_index=current_row,
                    existing_data={
                        'fluid': self._get_cell_value(ws, f'G{current_row}'),
                        'material_type': self._get_cell_value(ws, f'H{current_row}'),
                        'spec': self._get_cell_value(ws, f'I{current_row}'),
                        'grade': self._get_cell_value(ws, f'J{current_row}'),
                        'insulation': self._get_cell_value(ws, f'K{current_row}'),
                        'design_temp': self._get_cell_value(ws, f'L{current_row}'),
                        'design_pressure': self._get_cell_value(ws, f'M{current_row}'),
                        'operating_temp': self._get_cell_value(ws, f'N{current_row}'),
                        'operating_pressure': self._get_cell_value(ws, f'O{current_row}')
                    }
                )
                current_equipment.add_component(component_data)
                self.log_info(f"  - Component: {component_name} (row {current_row})")
            
            current_row += 1
        
        # Add the last equipment to dictionary
        if current_equipment and current_equipment.components:
            equipment_map[current_equipment.equipment_number] = current_equipment
        
        self.equipment_map = equipment_map
        total_components = sum(len(eq.components) for eq in equipment_map.values())
        self.log_info(f"📖 Read {len(equipment_map)} equipment items with {total_components} total components")
        self.close_workbook()
        return equipment_map
    
    # Updated methods to work with dictionary
    def get_equipment(self, equipment_number: str) -> Optional[Equipment]:
        """Get equipment by equipment number (O(1) lookup)"""
        return self.equipment_map.get(equipment_number)
    
    def get_all_equipment(self) -> List[Equipment]:
        """Get all equipment as list (for backward compatibility)"""
        return list(self.equipment_map.values())
    
    def equipment_exists(self, equipment_number: str) -> bool:
        """Check if equipment exists (O(1) check)"""
        return equipment_number in self.equipment_map
    
    def update_component_data(self, equipment_number: str, component_name: str, updates: Dict[str, any]) -> bool:
        """
        Update specific component data using setters
        Returns True if successful, False if equipment or component not found
        """
        equipment = self.equipment_map.get(equipment_number)
        if not equipment:
            self.log_warning(f"Equipment not found: {equipment_number}")
            return False
        
        component = equipment.get_component(component_name)
        if not component:
            self.log_warning(f"Component not found: {equipment_number} - {component_name}")
            return False
        
        try:
            component.update_existing_data(updates)
            self.log_info(f"Updated {equipment_number} - {component_name}: {updates}")
            return True
        except KeyError as e:
            self.log_error(f"Invalid data field {e} for {equipment_number} - {component_name}")
            return False
    
    def fill_empty_cells(self, equipment_number: str, component_name: str, default_values: Dict[str, any]) -> bool:
        """
        Fill empty cells for a specific component with default values
        """
        equipment = self.equipment_map.get(equipment_number)
        if not equipment:
            return False
        
        component = equipment.get_component(component_name)
        if not component:
            return False
        
        empty_fields = component.get_empty_data_fields()
        updates = {}
        
        for field in empty_fields:
            if field in default_values:
                updates[field] = default_values[field]
        
        if updates:
            component.update_existing_data(updates)
            self.log_info(f"Filled empty fields for {equipment_number} - {component_name}: {updates}")
            return True
        return False
    
    def add_new_equipment(self, equipment: Equipment) -> bool:
        """
        Add new equipment to the Dictionary
        Returns False if equipment with same number already exists
        """
        if equipment.equipment_number in self.equipment_map:
            self.log_warning(f"Equipment already exists: {equipment.equipment_number}")
            return False
        
        self.equipment_map[equipment.equipment_number] = equipment
        self.log_info(f"Added new equipment: {equipment.equipment_number}")
        return True
    
    def remove_equipment(self, equipment_number: str) -> bool:
        """
        Remove equipment from dictionary
        Returns True if removed, False if not found
        """
        if equipment_number in self.equipment_map:
            del self.equipment_map[equipment_number]
            self.log_info(f"Removed equipment: {equipment_number}")
            return True
        return False
    
    def get_equipment_by_pmt(self, pmt_number: str) -> List[Equipment]:
        """Get all equipment with specific PMT number"""
        return [eq for eq in self.equipment_map.values() if eq.pmt_number == pmt_number]
    
    def get_components_by_phase(self, phase: str) -> List[Component]:
        """Get all components with specific phase across all equipment"""
        components = []
        for equipment in self.equipment_map.values():
            for component in equipment.components:
                if component.phase == phase:
                    components.append(component)
        return components
    
    def save_to_excel(self, work_id: Optional[int] = None) -> bool:
        """
        Save the modified data back to Excel file
        """
        try:
            self.load_workbook()
            ws = self.wb['Masterfile']
            
            for equipment in self.equipment_map.values():
                for component in equipment.components:
                    if hasattr(component, 'row_index') and component.row_index:
                        row = component.row_index
                        # Update component data in Excel
                        ws[f'G{row}'] = component.get_existing_data_value('fluid')
                        ws[f'H{row}'] = component.get_existing_data_value('material_type')
                        ws[f'I{row}'] = component.get_existing_data_value('spec')
                        ws[f'J{row}'] = component.get_existing_data_value('grade')
                        ws[f'K{row}'] = component.get_existing_data_value('insulation')
                        ws[f'L{row}'] = component.get_existing_data_value('design_temp')
                        ws[f'M{row}'] = component.get_existing_data_value('design_pressure')
                        ws[f'N{row}'] = component.get_existing_data_value('operating_temp')
                        ws[f'O{row}'] = component.get_existing_data_value('operating_pressure')
            
            # Determine output path
            if work_id is None:
                os.makedirs(os.path.join(self.default_path, "default", "excel"), exist_ok=True)
                base, ext = os.path.splitext(self.file_path)
                path, base_name = os.path.split(base)
                path = "default/excel"
                output_path = os.path.join(self.default_path,path, f"{base_name}_modified{ext}")
            else:
                os.makedirs(os.path.join(self.default_path, f"user_{work_id}", "excel"), exist_ok=True)
                base, ext = os.path.splitext(self.file_path)
                path, base_name = os.path.split(base)
                path = f"{work_id}/excel"
                output_path = os.path.join(self.default_path, path, f"{work_id}_{base_name}_modified{ext}")
            # Save the workbook
            self.wb.save(output_path)
            self.log_info(f"✅ Excel file saved successfully: {output_path}")
            return True
            
        except Exception as e:
            self.log_error(f"❌ Error saving Excel file: {e}")
            return False
        finally:
            # Ensure workbook is closed in case of error
            if self._is_loaded:
                self.close_workbook()

    def add_timestamp(self, original_filename):
        # Get the current date and time
        timestamp_pattern = r"_\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}"
        new_base_name = re.sub(timestamp_pattern, '', original_filename)
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        # Construct the new filename
        new_filename = f"{new_base_name}_{timestamp}"
        
        return new_filename

    def save_to_excel_with_dict(self, equipment_dict: Dict[str, Equipment], user_id: Optional[int] = None) -> bool:
        """
        Save equipment dictionary data back to Excel file
        """
        try:
            # Load workbook if not loaded
            if not self._is_loaded:
                self.wb = load_workbook(self.file_path)
                self._is_loaded = True
            ws = self.wb['Masterfile']
            
            for equipment in equipment_dict.values():
                for component in equipment.components:
                    if hasattr(component, 'row_index') and component.row_index:
                        row = component.row_index
                        # Update component data in Excel
                        ws[f'G{row}'] = component.get_existing_data_value('fluid')
                        ws[f'H{row}'] = component.get_existing_data_value('material_type')
                        ws[f'I{row}'] = component.get_existing_data_value('spec')
                        ws[f'J{row}'] = component.get_existing_data_value('grade')
                        ws[f'K{row}'] = component.get_existing_data_value('insulation')
                        ws[f'L{row}'] = component.get_existing_data_value('design_temp')
                        ws[f'M{row}'] = component.get_existing_data_value('design_pressure')
                        ws[f'N{row}'] = component.get_existing_data_value('operating_temp')
                        ws[f'O{row}'] = component.get_existing_data_value('operating_pressure')
            
            # Determine output path
            if user_id is None:
                os.makedirs(os.path.join(self.default_path, "default", "excel"), exist_ok=True)
                base, ext = os.path.splitext(self.file_path)
                path, base_name = os.path.split(base)
                path = "default/excel"
                output_path = os.path.join(self.default_path, path, f"{base_name}_modified{ext}")
            else:
                os.makedirs(os.path.join(self.default_path, f"{user_id}", "excel", "updated"), exist_ok=True)
                base, ext = os.path.splitext(self.file_path)
                path, base_name = os.path.split(base)
                path = f"{user_id}/excel/updated"
                name_with_timestamp = self.add_timestamp(base_name.strip())
                output_path = os.path.join(self.default_path, path, f"{name_with_timestamp}{ext}")
            
            # Save the workbook
            self.wb.save(output_path)
            self.log_info(f"✅ Excel file saved successfully from dict: {output_path}")
            # Close workbook after saving
            self.close_workbook()
            return True
            
        except Exception as e:
            self.log_error(f"❌ Error saving Excel file from dict: {e}")
            return False
    
    def _get_cell_value(self, ws, cell_ref):
        """Safely get cell value handling merged cells"""
        try:
            return ws[cell_ref].value
        except:
            return None

    def equipment_to_json(self) -> str:
        """Convert equipment dict to JSON string"""
        import json
        equipment_list = [eq.to_dict() for eq in self.equipment_map.values()]
        return json.dumps(equipment_list, indent=2)
    
    def log_info(self, message: str) -> None:
        """Log info message to both console and UI"""
        logger.info(message)
        if self.log_callback:
            self.log_callback(f"ℹ️ {message}")

    def log_warning(self, message: str) -> None:
        """Log warning message to both console and UI"""
        logger.warning(message)
        if self.log_callback:
            self.log_callback(f"⚠️ {message}")

    def log_error(self, message: str) -> None:
        """Log error message to both console and UI"""
        logger.error(message)
        if self.log_callback:
            self.log_callback(f"❌ {message}")


if __name__ == "__main__":
    # Example usage
    extractor = ExcelManager("CaseStudy1Resources\\MasterFile _ IPETRO PLANT.xlsx")
    equipment_map = extractor.read_masterfile()
    
    # Print equipment
    print("Equipment Dict:")
    for equipment_number, equipment in equipment_map.items():
        print(f"Key: {equipment_number} -> {equipment}")
        for component in equipment.components:
            print(f"  - {component}")
            for key, value in component.existing_data.items():
                print(f"      {key}: {value}")
    
    #lookups
    specific_equipment = extractor.get_equipment("V-001")
    if specific_equipment:
        print(f"\nFound equipment: {specific_equipment}")
    
    # Check existence
    if extractor.equipment_exists("E-1002"):
        print("E-1002 exists!")
    else:
        print("E-1002 does not exist.")

    # Update component data
    extractor.update_component_data(
        equipment_number="V-002", 
        component_name="Shell", 
        updates={'fluid': 'Water', 'design_temp': 150}
    )
    
    # Save modified file
    #extractor.save_to_excel()
    #Save with user ID
    #extractor.save_to_excel(user_id=123)